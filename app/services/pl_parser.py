"""
P&L Excel parser service.

Parses a portfolio company's P&L Excel file (.xlsx) into structured monthly
data and aggregate summary KPIs.

Strategy
--------
- Load with openpyxl data_only=True to read cached formula results.
- Target the 'P&L' sheet (required). 'Cost' sheet is optional (headcount).
- 'Summary' sheet is optional (scenario label detection).
- Row identification: scan column A for known label substrings, stripping
  leading \\xa0 (non-breaking space) indentation characters.
- Month column detection: row 1 col C should be a cached datetime; subsequent
  months inferred by +1 month increments when their cells contain formula strings.
- Formula strings that were not cached (cell.value starts with '=') → None.
"""

import logging
from datetime import date, datetime
from typing import Optional

log = logging.getLogger(__name__)

# ── Helpers ──────────────────────────────────────────────────────────────────

def _safe_float(val) -> Optional[float]:
    """Return float or None. Rejects formula strings and non-numeric values."""
    if val is None:
        return None
    if isinstance(val, str):
        if val.startswith("="):
            return None  # uncached formula
        val = val.strip().rstrip("%")
        if not val:
            return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _clean_label(val) -> str:
    """Strip \\xa0 / whitespace and lowercase a cell's string value."""
    if val is None:
        return ""
    return str(val).replace("\xa0", " ").strip().lower()


def _cell_float(ws, row: int, col: int) -> Optional[float]:
    """Read a single cell as float-or-None."""
    return _safe_float(ws.cell(row=row, column=col).value)


# ── Month column detection ────────────────────────────────────────────────────

def _detect_month_columns(ws, start_col: int = 3, max_months: int = 24) -> list[tuple[int, date]]:
    """
    Return list of (col_index_1based, month_date) for the monthly data columns.

    Row 1 layout:  col A = label "P&L (Values in INR)", col B = "Total",
                   col C = first month date (cached datetime), cols D+ = formula strings.

    We read the first real datetime in row 1 (col C or later), then infer
    subsequent months by adding calendar months. We stop when we hit 2
    consecutive empty / non-date headers or reach max_months.
    """
    months: list[tuple[int, date]] = []
    first_date: Optional[date] = None
    consecutive_empty = 0

    def _add_months(d: date, n: int) -> date:
        """Add n calendar months to date d (stdlib-only, no dateutil)."""
        month = d.month - 1 + n
        year = d.year + month // 12
        month = month % 12 + 1
        import calendar
        day = min(d.day, calendar.monthrange(year, month)[1])
        return date(year, month, day)

    for col in range(start_col, start_col + max_months):
        cell_val = ws.cell(row=1, column=col).value

        # Try to extract a date from the cell
        month_date: Optional[date] = None
        if isinstance(cell_val, (datetime,)):
            month_date = cell_val.date() if hasattr(cell_val, "date") else cell_val
        elif isinstance(cell_val, date):
            month_date = cell_val
        elif cell_val is None or (isinstance(cell_val, str) and cell_val.startswith("=")):
            # Formula or empty — infer from first_date if we have one
            if first_date is not None:
                idx = len(months)
                month_date = _add_months(first_date, idx)
        # plain string date fallback
        elif isinstance(cell_val, str):
            try:
                month_date = datetime.strptime(cell_val.strip(), "%Y-%m-%d").date()
            except ValueError:
                pass

        if month_date is None:
            consecutive_empty += 1
            if consecutive_empty >= 2:
                break
            continue

        consecutive_empty = 0
        if first_date is None:
            first_date = month_date

        months.append((col, month_date))
        log.debug("MONTH_COL: col=%d → %s", col, month_date)

    log.info("PL_PARSER: detected %d month columns", len(months))
    return months


# ── Row label scanner ─────────────────────────────────────────────────────────

def _find_rows(ws) -> dict[str, int]:
    """
    Scan column A for known P&L row labels and return a mapping of
    field_name → 1-based row number.

    Only the first matching row is recorded for each field (prevents
    sub-total rows from overwriting the main row).
    """
    # Rules: (field_name, match_fn)
    # match_fn receives the cleaned, lowercase label string.
    # Order matters — more specific rules come first.

    found: dict[str, int] = {}

    for row in ws.iter_rows(min_col=1, max_col=1, values_only=False):
        cell = row[0]
        raw = cell.value
        if raw is None:
            continue
        label = _clean_label(raw)
        r = cell.row

        def _check(field: str, test: bool, _r: int = r) -> None:
            if test and field not in found:
                found[field] = _r
                log.debug("ROW_MAP: %s → row %d (label=%r)", field, _r, label)

        # Revenue block (top of sheet)
        _check("revenue",
               label == "revenue")
        _check("subscription_revenue",
               "subscription revenue" in label)
        _check("usage_credits",
               "usage credits" in label and "credit" in label)
        _check("revenue_growth_pct",
               "growth %" in label)

        # Cost of revenue
        _check("cost_of_revenue",
               label == "cost of revenue")

        # Gross profit
        _check("gross_profit",
               label == "gross profit")
        _check("gross_margin_pct",
               "gross margin" in label)

        # CAC / contribution — "spend on customer aquisition (incl. sales)"
        _check("cac_spend",
               "spend on customer" in label or
               ("customer" in label and ("cquisition" in label or "cquistion" in label)))
        _check("contribution_profit",
               label == "contribution profit")
        _check("contribution_margin_pct",
               "contribution margin" in label)

        # Operating cost lines
        _check("team_cost",
               "team cost" in label)
        _check("general_and_admin",
               "general and admin" in label)

        # Profit lines
        _check("ebitda",
               label == "ebitda")
        _check("ebit",
               label == "ebit")
        _check("net_profit",
               label == "net profit")

        # Cash section — "cash end" / "cash balance" / "ending cash" / "net cash"
        _check("cash_balance",
               label in ("cash end", "cash balance", "ending cash") or
               "ending cash" in label or "net cash" in label)

    log.info("PL_PARSER: row map found %d / 17 fields: %s", len(found), list(found.keys()))
    return found


def _find_cost_rows(ws_cost) -> dict[str, int]:
    """Scan Cost sheet col A for headcount rows."""
    found: dict[str, int] = {}
    for row in ws_cost.iter_rows(min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value is None:
            continue
        label = _clean_label(cell.value)
        r = cell.row
        if "team members" in label and "team_members" not in found:
            found["team_members"] = r
            log.debug("COST_ROW: team_members → row %d", r)
        if "engineering" in label and "founder" in label and "engineering_founders" not in found:
            found["engineering_founders"] = r
            log.debug("COST_ROW: engineering_founders → row %d", r)
    return found


# ── Scenario detection ────────────────────────────────────────────────────────

def _detect_scenario(wb) -> str:
    """
    Read the active scenario from Summary!B2 if the sheet exists.
    Falls back to "Base".
    """
    if "Summary" not in wb.sheetnames:
        return "Base"
    ws = wb["Summary"]
    val = ws["B2"].value
    if val and isinstance(val, str) and val.strip() in ("Base", "Best", "Worst"):
        return val.strip()
    return "Base"


# ── Summary computation ───────────────────────────────────────────────────────

def _compute_summary(months: list[dict]) -> dict:
    """Aggregate KPIs across all months."""

    def _sum(field: str) -> Optional[float]:
        vals = [m[field] for m in months if m.get(field) is not None]
        return round(sum(vals), 2) if vals else None

    def _last_nonnull(field: str) -> Optional[float]:
        for m in reversed(months):
            if m.get(field) is not None:
                return m[field]
        return None

    def _max_nonnull(field: str) -> Optional[float]:
        vals = [m[field] for m in months if m.get(field) is not None]
        return max(vals) if vals else None

    def _mean_nonnull(field: str) -> Optional[float]:
        vals = [m[field] for m in months if m.get(field) is not None]
        return round(sum(vals) / len(vals), 4) if vals else None

    def _first_positive_month(field: str) -> Optional[int]:
        for m in months:
            if m.get(field) is not None and m[field] > 0:
                return m["month_index"]
        return None

    # Final ARR in crores: last month's revenue * 12 / 10,000,000
    last_rev = _last_nonnull("revenue")
    final_arr_cr = round(last_rev * 12 / 10_000_000, 3) if last_rev else None

    return {
        "total_revenue": _sum("revenue"),
        "total_ebitda": _sum("ebitda"),
        "total_net_profit": _sum("net_profit"),
        "final_cash_balance": _last_nonnull("cash_balance"),
        "peak_team_members": _max_nonnull("team_members"),
        "final_arr_cr": final_arr_cr,
        "gross_margin_avg_pct": _mean_nonnull("gross_margin_pct"),
        "months_to_ebitda_positive": _first_positive_month("ebitda"),
        "months_to_net_positive": _first_positive_month("net_profit"),
        "runway_months": None,  # requires separate cash + burn — out of scope v1
    }


# ── Public entry point ────────────────────────────────────────────────────────

def parse_pl_excel(file_path: str) -> dict:
    """
    Parse a P&L Excel file and return a dict with keys:
        scenario  — "Base" / "Best" / "Worst"
        months    — list of monthly dicts (one per month column)
        summary   — aggregate KPI dict

    Raises ValueError if the file has no 'P&L' sheet.
    """
    import openpyxl

    log.info("PL_PARSER: loading %s", file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if "P&L" not in wb.sheetnames:
        raise ValueError(
            f"No 'P&L' sheet found in workbook. Available sheets: {wb.sheetnames}"
        )

    ws_pl = wb["P&L"]
    ws_cost = wb["Cost"] if "Cost" in wb.sheetnames else None

    # ── Detect scenario ───────────────────────────────────────────────────────
    scenario = _detect_scenario(wb)
    log.info("PL_PARSER: scenario=%s", scenario)

    # ── Detect month columns (row 1, starting col C = 3) ─────────────────────
    month_cols = _detect_month_columns(ws_pl, start_col=3, max_months=18)
    if not month_cols:
        raise ValueError("Could not detect any month columns in P&L row 1")

    # ── Find P&L row numbers ──────────────────────────────────────────────────
    pl_rows = _find_rows(ws_pl)

    # Cash balance fallback: row 52 is the standard position in this template
    # (labeled "Cash Balance" in the sheet's cash flow section)
    if "cash_balance" not in pl_rows:
        # Try row 52 directly
        label_52 = _clean_label(ws_pl.cell(row=52, column=1).value)
        if label_52:
            pl_rows["cash_balance"] = 52
            log.info("PL_PARSER: cash_balance fallback → row 52 (label=%r)", label_52)

    # ── Find Cost sheet row numbers (optional) ────────────────────────────────
    cost_rows: dict[str, int] = {}
    if ws_cost is not None:
        cost_rows = _find_cost_rows(ws_cost)

    # ── Extract monthly data ──────────────────────────────────────────────────
    months: list[dict] = []

    def _read_pl(field: str, col: int) -> Optional[float]:
        row_num = pl_rows.get(field)
        if row_num is None:
            return None
        return _cell_float(ws_pl, row_num, col)

    def _read_cost(field: str, col: int) -> Optional[float]:
        if ws_cost is None:
            return None
        row_num = cost_rows.get(field)
        if row_num is None:
            return None
        return _cell_float(ws_cost, row_num, col)

    for idx, (col, month_date) in enumerate(month_cols, start=1):
        month_entry: dict = {
            "month_index": idx,
            "month_date": month_date.isoformat(),
            "revenue": _read_pl("revenue", col),
            "subscription_revenue": _read_pl("subscription_revenue", col),
            "usage_credits": _read_pl("usage_credits", col),
            "revenue_growth_pct": _read_pl("revenue_growth_pct", col),
            "cost_of_revenue": _read_pl("cost_of_revenue", col),
            "gross_profit": _read_pl("gross_profit", col),
            "gross_margin_pct": _read_pl("gross_margin_pct", col),
            "cac_spend": _read_pl("cac_spend", col),
            "contribution_profit": _read_pl("contribution_profit", col),
            "contribution_margin_pct": _read_pl("contribution_margin_pct", col),
            "team_cost": _read_pl("team_cost", col),
            "general_and_admin": _read_pl("general_and_admin", col),
            "ebitda": _read_pl("ebitda", col),
            "ebit": _read_pl("ebit", col),
            "net_profit": _read_pl("net_profit", col),
            "cash_balance": _read_pl("cash_balance", col),
            "team_members": _read_cost("team_members", col),
            "engineering_founders": _read_cost("engineering_founders", col),
        }
        months.append(month_entry)

        if idx <= 3:
            log.info(
                "PL_MONTH[%d] %s: rev=%s gp=%s ebitda=%s net=%s team=%s",
                idx, month_date, month_entry["revenue"], month_entry["gross_profit"],
                month_entry["ebitda"], month_entry["net_profit"], month_entry["team_members"],
            )

    log.info("PL_PARSER: parsed %d months for scenario=%s", len(months), scenario)

    summary = _compute_summary(months)
    log.info("PL_PARSER: summary=%s", summary)

    return {
        "scenario": scenario,
        "months": months,
        "summary": summary,
    }


# ── Google Sheets pull mode ───────────────────────────────────────────────────

def parse_pl_from_sheets(sheets_url: str) -> dict:
    """
    Pull and parse a P&L Google Sheet using the Sheets API v4.

    The sheet must have the same structure as the Tessary P&L template:
      - Sheet named 'P&L' with computed values (Sheets evaluates formulas server-side)
      - Row 1 = headers: col A label, col B Total, cols C+ = month labels
      - Column A = row labels (same label conventions as parse_pl_excel)
      - Optional 'Cost' sheet for headcount rows
      - Optional 'Summary' sheet for scenario detection (B2)

    Uses UNFORMATTED_VALUE render option so every cell returns the numeric
    result of its formula — no formula strings to reject.

    Returns the same shape as parse_pl_excel():
        { scenario, months, summary }
    """
    import re as _re
    from app.config import settings
    from googleapiclient.discovery import build
    from google.oauth2 import service_account
    import os

    log.info("PL_SHEETS: pulling P&L from url=%s", sheets_url[:80])

    # ── Auth ──────────────────────────────────────────────────────────────────
    creds_file = settings.GOOGLE_CREDENTIALS_FILE.strip()
    if not os.path.exists(creds_file):
        raise ValueError("Google credentials file not configured or not found")

    creds = service_account.Credentials.from_service_account_file(
        creds_file,
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    svc = build("sheets", "v4", credentials=creds, cache_discovery=False)

    # ── Extract spreadsheet ID ────────────────────────────────────────────────
    match = _re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", sheets_url)
    if not match:
        raise ValueError(f"Could not extract spreadsheet ID from URL: {sheets_url}")
    spreadsheet_id = match.group(1)
    log.info("PL_SHEETS: spreadsheet_id=%s", spreadsheet_id)

    def _get_range(range_name: str):
        result = svc.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueRenderOption="UNFORMATTED_VALUE",
            dateTimeRenderOption="FORMATTED_STRING",
        ).execute()
        return result.get("values", [])

    # ── Verify sheet exists ───────────────────────────────────────────────────
    meta = svc.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheet_names = {s["properties"]["title"] for s in meta.get("sheets", [])}
    log.info("PL_SHEETS: available sheets=%s", sorted(sheet_names))
    if "P&L" not in sheet_names:
        raise ValueError(
            f"No 'P&L' sheet found. Available: {sorted(sheet_names)}"
        )

    # ── Scenario from Summary!B2 ──────────────────────────────────────────────
    scenario = "Base"
    if "Summary" in sheet_names:
        try:
            summary_rows = _get_range("'Summary'!B2")
            val = (summary_rows[0][0] if summary_rows and summary_rows[0] else "")
            if str(val).strip() in ("Base", "Best", "Worst"):
                scenario = str(val).strip()
        except Exception as e:
            log.warning("PL_SHEETS: could not read scenario: %s", e)
    log.info("PL_SHEETS: scenario=%s", scenario)

    # ── Read P&L sheet ─────────────────────────────────────────────────────────
    pl_rows = _get_range("'P&L'!A1:T55")
    log.info("PL_SHEETS: received %d rows from P&L", len(pl_rows))

    # ── Month columns from row 1 ──────────────────────────────────────────────
    # Row 1 format: ['P&L (Values in INR)', 'Total', ' Apr 26', ' May 26', ...]
    # Sheets returns these as formatted strings (month labels), not dates.
    # We parse them and fall back to sequential increments.
    if not pl_rows:
        raise ValueError("P&L sheet returned no data")

    header_row = pl_rows[0]  # 0-indexed list

    from datetime import date as _date
    import calendar as _cal

    def _parse_month_label(s) -> Optional[_date]:
        """Parse labels like ' Apr 26', 'Apr 2026', 'April 26', '2026-04-01'."""
        if not s:
            return None
        s = str(s).strip()
        for fmt in ("%b %y", "%b %Y", "%B %y", "%B %Y", "%Y-%m-%d"):
            try:
                from datetime import datetime as _dt
                d = _dt.strptime(s, fmt).date()
                return d.replace(day=1)
            except ValueError:
                continue
        return None

    # Cols C onward = index 2+ in the list
    month_dates: list[_date] = []
    first_date: Optional[_date] = None
    for col_idx in range(2, min(len(header_row), 22)):  # up to 20 month cols
        val = header_row[col_idx]
        d = _parse_month_label(val)
        if d is None and first_date is not None:
            # Infer: +N months from first_date
            n = len(month_dates)
            m = first_date.month - 1 + n
            yr = first_date.year + m // 12
            mo = m % 12 + 1
            d = _date(yr, mo, 1)
        if d is not None:
            if first_date is None:
                first_date = d
            month_dates.append(d)

    log.info("PL_SHEETS: detected %d month columns", len(month_dates))
    if not month_dates:
        raise ValueError("Could not detect month columns in P&L row 1")

    # ── Build row-label → row-index map ───────────────────────────────────────
    # pl_rows is 0-indexed; row labels are in column A (index 0)
    row_map: dict[str, int] = {}  # field_name → 0-based row index in pl_rows

    for ri, row in enumerate(pl_rows):
        if not row:
            continue
        label = _clean_label(row[0])
        if not label:
            continue

        def _reg(field: str, test: bool, _ri: int = ri) -> None:
            if test and field not in row_map:
                row_map[field] = _ri

        _reg("revenue",                  label == "revenue")
        _reg("subscription_revenue",     "subscription revenue" in label)
        _reg("usage_credits",            "usage credits" in label and "credit" in label)
        _reg("revenue_growth_pct",       "growth %" in label)
        _reg("cost_of_revenue",          label == "cost of revenue")
        _reg("gross_profit",             label == "gross profit")
        _reg("gross_margin_pct",         "gross margin" in label)
        _reg("cac_spend",                "spend on customer" in label or
                                         ("customer" in label and "cquisition" in label))
        _reg("contribution_profit",      label == "contribution profit")
        _reg("contribution_margin_pct",  "contribution margin" in label)
        _reg("team_cost",                "team cost" in label)
        _reg("general_and_admin",        "general and admin" in label)
        _reg("ebitda",                   label == "ebitda")
        _reg("ebit",                     label == "ebit")
        _reg("net_profit",               label == "net profit")
        _reg("cash_balance",             label in ("cash end", "cash balance", "ending cash") or
                                         "ending cash" in label or "net cash" in label)

    log.info("PL_SHEETS: row map found %d / 17 fields: %s",
             len(row_map), list(row_map.keys()))

    # ── Headcount from Cost sheet ─────────────────────────────────────────────
    cost_row_map: dict[str, int] = {}
    if "Cost" in sheet_names:
        try:
            cost_rows = _get_range("'Cost'!A1:T30")
            for ri, row in enumerate(cost_rows):
                if not row:
                    continue
                label = _clean_label(row[0])
                if "team members" in label and "team_members" not in cost_row_map:
                    cost_row_map["team_members"] = ri
                if "engineering" in label and "founder" in label and "engineering_founders" not in cost_row_map:
                    cost_row_map["engineering_founders"] = ri
        except Exception as e:
            log.warning("PL_SHEETS: could not read Cost sheet: %s", e)

    # ── Extract monthly values ────────────────────────────────────────────────
    # Sheets columns C-T map to list indices 2-19 → month_dates index 0-17
    def _get_val(rows: list, row_idx: int, col_offset: int) -> Optional[float]:
        """row_idx is 0-based index into `rows`; col_offset is 0-based (0=colA)."""
        try:
            row = rows[row_idx]
            val = row[col_offset] if col_offset < len(row) else None
            return _safe_float(val)
        except (IndexError, TypeError):
            return None

    # Cost rows data (already fetched above as cost_rows)
    cost_data: list = []
    if "Cost" in sheet_names:
        try:
            cost_data = _get_range("'Cost'!A1:T30")
        except Exception:
            pass

    months_out: list[dict] = []
    for idx, month_date in enumerate(month_dates, start=1):
        col_offset = idx + 1   # col A=0, B=1, C=2 → month 1 = offset 2

        def _pl(field: str) -> Optional[float]:
            ri = row_map.get(field)
            if ri is None:
                return None
            return _get_val(pl_rows, ri, col_offset)

        def _cost(field: str) -> Optional[float]:
            ri = cost_row_map.get(field)
            if ri is None:
                return None
            return _get_val(cost_data, ri, col_offset)

        month_entry = {
            "month_index": idx,
            "month_date": month_date.isoformat(),
            "revenue": _pl("revenue"),
            "subscription_revenue": _pl("subscription_revenue"),
            "usage_credits": _pl("usage_credits"),
            "revenue_growth_pct": _pl("revenue_growth_pct"),
            "cost_of_revenue": _pl("cost_of_revenue"),
            "gross_profit": _pl("gross_profit"),
            "gross_margin_pct": _pl("gross_margin_pct"),
            "cac_spend": _pl("cac_spend"),
            "contribution_profit": _pl("contribution_profit"),
            "contribution_margin_pct": _pl("contribution_margin_pct"),
            "team_cost": _pl("team_cost"),
            "general_and_admin": _pl("general_and_admin"),
            "ebitda": _pl("ebitda"),
            "ebit": _pl("ebit"),
            "net_profit": _pl("net_profit"),
            "cash_balance": _pl("cash_balance"),
            "team_members": _cost("team_members"),
            "engineering_founders": _cost("engineering_founders"),
        }
        months_out.append(month_entry)

        if idx <= 3:
            log.info(
                "PL_SHEETS_MONTH[%d] %s: rev=%s gp=%s ebitda=%s net=%s",
                idx, month_date,
                month_entry["revenue"], month_entry["gross_profit"],
                month_entry["ebitda"], month_entry["net_profit"],
            )

    log.info("PL_SHEETS: parsed %d months for scenario=%s", len(months_out), scenario)
    summary = _compute_summary(months_out)
    log.info("PL_SHEETS: summary=%s", summary)

    return {
        "scenario": scenario,
        "months": months_out,
        "summary": summary,
    }
